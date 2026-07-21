#!/usr/bin/env python3
"""Regenerate the age-restricted keyword brain from the canonical Gopher iQ xlsx.

Source of truth (owner-maintained):
  Documentation/Dashboard/Gopher iQ/Age-Restricted Key Words/
    Age_Restricted_Tobacco_Keywords.xlsx        (sheet "Keywords", col "Keyword")
    Age_Restricted_Alcohol_Keywords.xlsx        (sheet "Keywords", col "Keyword" = idx 3)
    Popular_Beer_Wine_Liquor_Keywords.xlsx      (sheet "Alcohol Keywords", col "Keyword" = idx 1)
    Gopher_iQ_Tobacco_Nicotine_Taxonomy_with_Flavors.xlsx  (sheet "Engine Import")

All four are UNIONED. The taxonomy workbook SUPPLEMENTS the older tobacco file, it
does not replace it: the older file carries 29 terms the taxonomy lacks, including
bare "nicotine", the vape brand "blu", common misspellings (malboro, nicotene) and
the intent phrases ("smoke run", "need cigs", "pick up dip"). Dropping it would be
a silent coverage regression.

Output (do NOT hand-edit):
  Final/assets/js/gopher-age-keywords.js   ->  window.GopherAgeKeywords = [...]

Consumed by findAgeRestrictedKeyword() in:
  Final/gopher-request.html, Final/gopher-connect.html,
  _prototypes/Request/gopher-request-flow.html

Run:  python3 docs/handoff/regen-age-keywords.py
Requires: openpyxl.  When the owner adds brands to the xlsx, re-run this.
"""
import openpyxl, json, os

HERE = os.path.dirname(os.path.abspath(__file__))              # .../Code/docs/handoff
REPO = os.path.dirname(os.path.dirname(HERE))                  # .../Code
# Dashboard lives two levels above the Code repo: .../Documentation/Dashboard
SRC  = os.path.join(REPO, "..", "..", "Dashboard", "Gopher iQ", "Age-Restricted Key Words")
OUT  = os.path.join(REPO, "Final", "assets", "js", "gopher-age-keywords.js")

def norm(s): return " ".join(str(s).strip().lower().split())

# The alcohol xlsx tags each row with a Keyword_Type. Its own Summary sheet says:
# "Flag High confidence matches automatically; queue Medium confidence slang/generic
# matches for review." The "Slang or generic keyword" rows are everyday words
# (bottle, can, shot, pint, cold one, ...) that would false-positive on ordinary
# grocery orders, so we EXCLUDE them from the auto-flag brain per that guidance.
ALCOHOL_EXCLUDE_TYPES = {"slang or generic keyword"}

# Everyday-word collisions that survive in the brand-only lists (tobacco/beer files
# have no type column). "carton of cigarettes" stays; bare "carton" (of eggs/milk) goes.
GENERIC_STOPLIST = {
    "bottle", "bottles", "can", "cans", "carton", "cartons", "pack", "packs",
    "case", "cases", "box", "boxes", "shot", "shots", "pint", "pints",
    "glass", "glasses", "cold one", "cold ones",
}

# Bare words the taxonomy lists as tobacco terms that are ALSO ordinary delivery
# vocabulary. The matcher is whole-word, so these do not bleed inside other words --
# but on their own they would flag routine orders and force ID verification on them:
#   juice -> "orange juice"        pipe   -> "pvc pipe"        wraps -> "chicken wraps"
#   dip   -> "french onion dip"    cones  -> "ice cream cones"  heater -> "space heater"
#   ends/true/west/zone/square(s)/punch/matches/loop/grinder/native/woods/chew/burn/drag
# This is the workbook's OWN instruction, not our invention. Its Implementation Guide:
#   "Do not classify a request solely from weak ambiguous words. Require product,
#    store, brand, delivery, purchase, or age-verification context."
# and it names dip / pod / juice / pack / smoke as needing stricter thresholds.
#
# Multi-word forms are unaffected -- "marlboro reds" stays even though bare "reds" goes,
# "blunt wraps" stays while bare "wraps" goes. The terms remain in the canonical xlsx;
# a production context-aware matcher should use them WITH a co-occurring signal.
AMBIGUOUS_REQUIRE_CONTEXT = {
    # everyday nouns / verbs
    "ace", "acid", "azure", "burn", "chew", "cones", "cougar", "dart", "dip",
    "drag", "dutch", "ends", "fogger", "glo", "golds", "grinder", "heater",
    "heaters", "husky", "juice", "kayak", "loop", "matchbook", "matches",
    "mouthpiece", "native", "neo", "pipe", "ports", "pouch", "pouches", "punch",
    "square", "squares", "true", "volt", "west", "woods", "wraps", "zone",
    # cigarette brands that are also common proper nouns. "salem" matters most here:
    # the matcher treats the hyphen in "Winston-Salem" as a boundary, so every
    # delivery naming an NC city Gopher actually serves would flag as tobacco.
    "capri", "carlton", "fantasia", "kent", "lucy", "merit", "misty", "optimo",
    "salem", "seneca", "spaceman", "vantage",
    # baby-food brand that collides with a chewing-tobacco brand
    "beech-nut",
    # PRE-EXISTING false positives, from the older tobacco file rather than the new
    # taxonomy — both confirmed firing before this supplement was added:
    #   "winston" flagged "deliver to Winston-Salem NC"  (an NC city Gopher serves;
    #             the hyphen counts as a word boundary, so both halves matched)
    #   "game"    flagged "board game for a birthday"
    "winston", "game",
}

def collect(fname, sheet, col, acc, type_col=None):
    wb = openpyxl.load_workbook(os.path.join(SRC, fname), read_only=True, data_only=True)
    ws = wb[sheet]
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:  # header
            continue
        if not (row and len(row) > col and row[col] is not None):
            continue
        if type_col is not None and len(row) > type_col and row[type_col] is not None:
            if norm(row[type_col]) in ALCOHOL_EXCLUDE_TYPES:
                continue
        s = norm(row[col])
        if s and len(s) >= 2:
            acc.add(s)
    wb.close()

def collect_taxonomy(fname, acc):
    """Gopher_iQ_Tobacco_Nicotine_Taxonomy_with_Flavors.xlsx -> sheet "Engine Import".
    Columns: keyword | normalized_category | term_type | age_restricted | confidence.

    Two filters, both taken from the workbook itself rather than chosen by us:

    1. age_restricted == "Yes" only. The sheet already separates auto-flag terms
       (711) from "Context Dependent" ones (522, nearly all Flavor rows). Its
       Implementation Guide: "A flavor alone should not automatically trigger age
       restriction. Combine it with a tobacco/nicotine product, brand, store,
       purchase, delivery, or usage signal." Flat-listing flavours would flag
       "mint gum", "cherry tomatoes", "vanilla extract" and "blue raspberry slushie".

    2. Skip rows containing [ or ]. The "Pattern Template" rows are slot patterns,
       not literal keywords -- "pack of [brand]", "[flavor] vape", "[size] cigar".
       Taken literally they would only ever match the bracket text itself.
    """
    wb = openpyxl.load_workbook(os.path.join(SRC, fname), read_only=True, data_only=True)
    ws = wb["Engine Import"]
    kept = skipped_ctx = skipped_tpl = 0
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0 or not row or row[0] is None:
            continue
        term = norm(row[0])
        age  = norm(row[3]) if len(row) > 3 and row[3] is not None else ""
        if age != "yes":
            skipped_ctx += 1;  continue
        if "[" in term or "]" in term:
            skipped_tpl += 1;  continue
        if len(term) >= 2:
            acc.add(term); kept += 1
    wb.close()
    print("  taxonomy: kept %d | skipped %d context-dependent | skipped %d pattern templates"
          % (kept, skipped_ctx, skipped_tpl))

def main():
    kws = set()
    collect("Age_Restricted_Tobacco_Keywords.xlsx",   "Keywords",         0, kws)
    collect("Age_Restricted_Alcohol_Keywords.xlsx",   "Keywords",         3, kws, type_col=4)
    collect("Popular_Beer_Wine_Liquor_Keywords.xlsx", "Alcohol Keywords", 1, kws)
    collect_taxonomy("Gopher_iQ_Tobacco_Nicotine_Taxonomy_with_Flavors.xlsx", kws)
    kws -= GENERIC_STOPLIST
    dropped = sorted(kws & AMBIGUOUS_REQUIRE_CONTEXT)
    kws -= AMBIGUOUS_REQUIRE_CONTEXT
    if dropped:
        print("  held back as context-required (%d): %s" % (len(dropped), ", ".join(dropped)))
    arr = sorted(kws)
    hdr = ("/* AUTO-GENERATED — Gopher iQ age-restricted keyword brain.\n"
           "   Source of truth: Documentation/Dashboard/Gopher iQ/Age-Restricted Key Words/\n"
           "     Age_Restricted_Tobacco_Keywords.xlsx (col Keyword)\n"
           "     Age_Restricted_Alcohol_Keywords.xlsx (col Keyword)\n"
           "     Popular_Beer_Wine_Liquor_Keywords.xlsx (col Keyword)\n"
           "     Gopher_iQ_Tobacco_Nicotine_Taxonomy_with_Flavors.xlsx (sheet Engine Import)\n"
           "   All four are unioned; the taxonomy SUPPLEMENTS the older tobacco file.\n"
           "   AUTO-FLAG SUBSET ONLY. Two groups are deliberately excluded and are NOT\n"
           "   missing data — they need a context-aware matcher, per the taxonomy's own\n"
           "   Implementation Guide:\n"
           "     * age_restricted = 'Context Dependent' (flavours: mint, cherry, vanilla…)\n"
           "       'A flavor alone should not automatically trigger age restriction.'\n"
           "     * bare ambiguous words that are ordinary delivery vocabulary\n"
           "       (juice, pipe, wraps, dip, cones, salem, …) — see AMBIGUOUS_REQUIRE_CONTEXT\n"
           "       in the generator. Multi-word forms are kept ('blunt wraps', 'marlboro reds').\n"
           "   %d unique keywords. Do NOT hand-edit — regenerate: python3 docs/handoff/regen-age-keywords.py\n"
           "   Consumed by findAgeRestrictedKeyword() in the Request flow + Final apps. */\n" % len(arr))
    js = hdr + "window.GopherAgeKeywords=" + json.dumps(arr, ensure_ascii=False, separators=(",", ":")) + ";\n"
    with open(OUT, "w", encoding="utf-8") as f:
        f.write(js)
    print("wrote %s | %d keywords | %d bytes" % (OUT, len(arr), len(js)))

if __name__ == "__main__":
    main()
